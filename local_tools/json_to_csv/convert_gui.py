import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk
import json
import os
from datetime import datetime

try:
    import pandas as pd
    import pytz
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass  # handled at runtime

# --- Config ---
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR  = os.path.join(BASE_DIR, 'input')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
os.makedirs(INPUT_DIR,  exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

TZ = pytz.timezone('Asia/Taipei')

# ── JSON Parsing ──────────────────────────────────────────────────────────────

def ts_to_str(ms):
    """Unix ms → 台灣時間字串"""
    if not ms:
        return ''
    try:
        return datetime.fromtimestamp(ms / 1000.0, TZ).strftime('%Y-%m-%d %H:%M')
    except Exception:
        return ''

def order_type_zh(raw):
    mapping = {'dine_in': '內用', 'takeout': '外帶', 'delivery': '外送', 'pick_up': '自取'}
    return mapping.get(raw, raw)

def parse_json_to_dataframes(file_path):
    """
    解析 Eats365 JSON，回傳兩個 DataFrame：
      df_orders  → 訂單總覽（一筆訂單一列）
      df_items   → 商品明細（一個品項一列）
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        raw = json.load(f)

    # 找到 pay_load.data（或任何最大的 list）
    orders = raw.get('pay_load', {}).get('data', [])
    if not orders:
        raise ValueError("找不到訂單資料（pay_load.data 為空）")

    order_rows = []
    item_rows  = []

    for order in orders:
        # ── 基本欄位 ──
        created_at     = ts_to_str(order.get('created_at'))
        short_code     = order.get('short_code', '')
        order_id       = order.get('id', '')
        status_raw     = order.get('status', '')
        status         = '已完成' if status_raw == 'COMPLETED' else status_raw
        o_type         = order_type_zh(order.get('type', ''))
        party_size     = order.get('party_size', 0)
        total          = order.get('total_price', 0)
        note           = order.get('note', '')
        receipt_number = order.get('receipt_number', '')

        # ── 付款方式 ──
        payments = order.get('payments', [])
        payment_method = payments[0].get('name', '') if payments else ''
        payment_amount = payments[0].get('amount', 0) if payments else 0

        # ── 客戶資訊 ──
        ship = order.get('shipping_information', {})
        receiver = ship.get('receiver', {}) if isinstance(ship, dict) else {}
        customer_name  = receiver.get('first_name', '')
        customer_phone = receiver.get('phone_number', '')

        # ── 發票 ──
        invoice_list   = order.get('invoice_list', [])
        invoice_number = invoice_list[0].get('invoice_number', '') if invoice_list else ''

        # ── 折扣 ──
        discounts = order.get('discounts', [])
        discount_names   = '、'.join([d.get('name', {}).get('default', '') for d in discounts if isinstance(d, dict)])
        discount_amounts = sum(d.get('discounted_amount', 0) for d in discounts if isinstance(d, dict))

        order_rows.append({
            '日期時間':   created_at,
            '收據號碼':   receipt_number,
            '短號':       short_code,
            '狀態':       status,
            '訂單類型':   o_type,
            '人數':       party_size,
            '金額':       total,
            '折扣金額':   discount_amounts,
            '折扣說明':   discount_names,
            '付款方式':   payment_method,
            '付款金額':   payment_amount,
            '客戶姓名':   customer_name,
            '客戶電話':   customer_phone,
            '發票號碼':   invoice_number,
            '備註':       note,
            '訂單ID':     order_id,
        })

        # ── 商品明細 ──
        line_items = order.get('line_items', [])
        for item in line_items:
            item_name  = item.get('product_name', {}).get('default', '')
            sku        = item.get('product_code', '')
            qty        = item.get('quantity', 0)
            unit_price = item.get('price', 0)
            subtotal   = item.get('price_total', 0)
            is_combo   = item.get('is_combo', False)
            mods       = item.get('modifiers', [])
            options    = '、'.join([m.get('value_name', {}).get('zh_TW', '') or
                                    m.get('value_name', {}).get('default', '')
                                    for m in mods if isinstance(m, dict)])

            item_rows.append({
                '日期時間': created_at,
                '短號':     short_code,
                '訂單類型': o_type,
                '商品名稱': item_name,
                'SKU':      sku,
                '數量':     qty,
                '單價':     unit_price,
                '小計':     subtotal,
                '選項':     options,
                '是否套餐': '是' if is_combo else '',
            })

            # 套餐子商品
            for b in item.get('bundled', []):
                b_name = b.get('product_name', {}).get('default', '') or \
                         b.get('product_name', {}).get('zh_TW', '')
                b_sku  = b.get('product_code', '')
                b_qty  = b.get('quantity', 0) * qty
                b_mods = b.get('modifiers', [])
                b_opts = '、'.join([m.get('value_name', {}).get('zh_TW', '') or
                                    m.get('value_name', {}).get('default', '')
                                    for m in b_mods if isinstance(m, dict)])
                item_rows.append({
                    '日期時間': created_at,
                    '短號':     short_code,
                    '訂單類型': o_type,
                    '商品名稱': f'  └ {b_name}',
                    'SKU':      b_sku,
                    '數量':     b_qty,
                    '單價':     b.get('price', 0),
                    '小計':     b.get('price_total', 0),
                    '選項':     b_opts,
                    '是否套餐': '套餐子項',
                })

    df_orders = pd.DataFrame(order_rows)
    df_items  = pd.DataFrame(item_rows)
    return df_orders, df_items


# ── Excel Styling ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
BORDER_SIDE  = Side(style='thin', color='BFBFBF')
CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal='left',   vertical='center', wrap_text=False)


def style_sheet(ws):
    """套用標題列樣式 + 斑馬紋 + 自動欄寬"""
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = CELL_BORDER

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border    = CELL_BORDER
            cell.alignment = LEFT_ALIGN

    # 自動欄寬
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    ws.freeze_panes = 'A2'


def save_excel(df_orders, df_items, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_orders.to_excel(writer, sheet_name='訂單總覽', index=False)
        df_items.to_excel(writer,  sheet_name='商品明細', index=False)

    # 套用樣式
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        style_sheet(wb[sheet_name])
    wb.save(output_path)


# ── GUI ───────────────────────────────────────────────────────────────────────

class JsonReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("滾麵 每日結帳 JSON → Excel 報表")
        self.root.geometry("640x440")
        self.root.resizable(False, False)

        tk.Label(root, text="每日結帳 JSON 轉 Excel 報表", font=("Arial", 18, "bold")).pack(pady=10)

        instruction_text = (
            f"1. 請把每日結帳 .txt / .json 放入 '{os.path.basename(INPUT_DIR)}' 資料夾。\n"
            f"2. 按下方按鈕，自動產生兩個工作表：\n"
            f"   • 訂單總覽（日期、類型、金額、折扣、付款、客戶）\n"
            f"   • 商品明細（商品名稱、SKU、數量、單價）\n"
            f"3. Excel 輸出至 '{os.path.basename(OUTPUT_DIR)}' 資料夾。"
        )
        tk.Label(root, text=instruction_text, justify="left", font=("Arial", 11)).pack(pady=6, padx=20, anchor="w")

        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=10, pady=4)

        self.btn = tk.Button(
            root, text="🚀  開始產生報表",
            command=self.run,
            font=("Arial", 14, "bold"), bg="#1F78B4", fg="white", height=2
        )
        self.btn.pack(pady=8, fill='x', padx=50)

        tk.Label(root, text="執行日誌 (Log):", font=("Arial", 10, "bold")).pack(anchor='w', padx=12)
        self.log_area = scrolledtext.ScrolledText(root, height=9, state='disabled', font=("Courier", 10))
        self.log_area.pack(fill='both', expand=True, padx=10, pady=5)

        self.log(f"準備就緒。請將 JSON / TXT 放入: {INPUT_DIR}")

    def log(self, msg):
        self.log_area.config(state='normal')
        self.log_area.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_area.see(tk.END)
        self.log_area.config(state='disabled')
        self.root.update()

    def run(self):
        self.btn.config(state='disabled')
        self.log("--- 開始掃描 ---")

        files = [f for f in os.listdir(INPUT_DIR) if f.lower().endswith(('.json', '.txt'))]
        if not files:
            self.log("❌ 找不到檔案！請確認 input 資料夾有 .json 或 .txt。")
            messagebox.showwarning("找不到檔案", "請放入 .json 或 .txt 檔案後再試。")
            self.btn.config(state='normal')
            return

        self.log(f"找到 {len(files)} 個檔案，開始轉換...")
        ok = 0
        for filename in files:
            input_path  = os.path.join(INPUT_DIR, filename)
            out_name    = os.path.splitext(filename)[0] + "_報表.xlsx"
            output_path = os.path.join(OUTPUT_DIR, out_name)
            try:
                self.log(f"正在解析: {filename} ...")
                df_orders, df_items = parse_json_to_dataframes(input_path)
                save_excel(df_orders, df_items, output_path)
                self.log(f"  ↳ 訂單 {len(df_orders)} 筆 / 品項 {len(df_items)} 列")
                self.log(f"✅ 完成: {out_name}")
                ok += 1
            except Exception as e:
                self.log(f"❌ 失敗 {filename}: {e}")

        self.log(f"--- 完成 {ok}/{len(files)} ---")
        messagebox.showinfo("完成", f"成功產生 {ok} 份報表！\n位置：{OUTPUT_DIR}")
        self.btn.config(state='normal')


if __name__ == "__main__":
    try:
        root = tk.Tk()
        app  = JsonReportApp(root)
        root.mainloop()
    except Exception as e:
        print(f"Error: {e}")
        input("Press Enter to exit...")
