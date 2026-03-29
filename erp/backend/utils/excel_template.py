"""
動態產生匯入用 Excel 範本（openpyxl）
- 標頭：必填橘色，選填深灰
- 範例列：低飽和藍色背景，灰色斜體
- 下拉驗證：從隱藏 _Lists sheet 取值（動態資料）
- 凍結第一列
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ─── 樣式 ────────────────────────────────────────────────────────────────────

REQ_FILL   = PatternFill("solid", fgColor="E85D04")   # 橘，必填
OPT_FILL   = PatternFill("solid", fgColor="374151")   # 深灰，選填
EX_FILL    = PatternFill("solid", fgColor="1A2535")   # 範例列底色
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
EX_FONT    = Font(color="6B7280", italic=True, size=10)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header(ws, headers):
    """headers: list of (label, required, width)"""
    ws.row_dimensions[1].height = 26
    for col, (label, req, width) in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill = REQ_FILL if req else OPT_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = width


def _example(ws, values):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=2, column=col, value=val)
        c.fill = EX_FILL
        c.font = EX_FONT


def _dropdown(ws, formula1, col_letter, start=2, end=2000):
    dv = DataValidation(
        type="list", formula1=formula1,
        allowBlank=True, showErrorMessage=True,
        error="請從下拉選單選擇", errorTitle="無效輸入"
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start}:{col_letter}{end}")


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 品項範本 ─────────────────────────────────────────────────────────────────

def build_items_template(vendors: list, groups: list) -> bytes:
    """
    vendors: [{'name': str}, ...]
    groups:  [{'name': str}, ...]
    欄位：名稱*, 單位*, 供應商, 盤點群組, 安全庫存, 參考價格
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "品項匯入"
    ws.freeze_panes = "A2"

    _header(ws, [
        ("名稱 *",  True,  18),
        ("單位 *",  True,  10),
        ("供應商",  False, 16),
        ("盤點群組", False, 16),
        ("安全庫存", False, 12),
        ("參考價格", False, 12),
    ])
    _example(ws, [
        "白蘿蔔", "條",
        vendors[0]["name"] if vendors else "",
        groups[0]["name"]  if groups  else "",
        5, 30,
    ])

    lists = wb.create_sheet("_Lists")
    lists.sheet_state = "hidden"

    if vendors:
        for i, v in enumerate(vendors, 1):
            lists.cell(row=i, column=1, value=v["name"])
        _dropdown(ws, f"_Lists!$A$1:$A${len(vendors)}", "C")

    if groups:
        for i, g in enumerate(groups, 1):
            lists.cell(row=i, column=2, value=g["name"])
        _dropdown(ws, f"_Lists!$B$1:$B${len(groups)}", "D")

    return _to_bytes(wb)


# ─── 供應商範本 ───────────────────────────────────────────────────────────────

PAYMENT_TERMS = ["先收款", "現付", "後收款", "週結", "月結"]
YN = ["是", "否"]

def build_vendors_template(payment_methods: list, categories: list) -> bytes:
    """
    payment_methods: [{'name': str}, ...]
    categories:      [{'name': str}, ...]
    欄位：名稱*, 聯絡人, 電話, LINE ID, 付款方式, 付款條件, 金流科目,
          到期提醒天數, 匯款帳號, 戶名, 到貨天數, 免運門檻, 出現叫貨系統, 固定排程,
          叫貨星期, 截單時間, 休息日星期, 國定假日休息, 備注
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "供應商匯入"
    ws.freeze_panes = "A2"

    _header(ws, [
        ("名稱 *",       True,  16),
        ("聯絡人",       False, 12),
        ("電話",         False, 14),
        ("LINE ID",      False, 14),
        ("付款方式",     False, 14),
        ("付款條件",     False, 14),
        ("金流科目",     False, 16),
        ("到期提醒(天)", False, 14),
        ("匯款帳號",     False, 20),
        ("戶名",         False, 14),
        ("到貨天數",     False, 12),
        ("免運門檻",     False, 12),
        ("出現叫貨系統", False, 14),
        ("固定排程",     False, 12),
        ("叫貨星期",     False, 16),   # O: 逗號分隔 1-7，1=一…7=日
        ("截單時間",     False, 12),   # P: HH:MM
        ("休息日星期",   False, 16),   # Q: 逗號分隔 1-7
        ("國定假日休息", False, 14),   # R: 是/否
        ("備注",         False, 20),   # S
    ])

    pm_name = payment_methods[0]["name"] if payment_methods else "現金"
    cat_name = categories[0]["name"] if categories else ""
    _example(ws, [
        "點線麵", "王小明", "0912345678", "",
        pm_name, "月結", cat_name,
        5, "", "", 1, "", "是", "是",
        "1,3,5", "14:00", "6,7", "是", "",
    ])

    lists = wb.create_sheet("_Lists")
    lists.sheet_state = "hidden"

    # A: 付款方式（動態）
    for i, m in enumerate(payment_methods, 1):
        lists.cell(row=i, column=1, value=m["name"])
    if payment_methods:
        _dropdown(ws, f"_Lists!$A$1:$A${len(payment_methods)}", "E")

    # B: 付款條件（靜態）
    for i, t in enumerate(PAYMENT_TERMS, 1):
        lists.cell(row=i, column=2, value=t)
    _dropdown(ws, f"_Lists!$B$1:$B${len(PAYMENT_TERMS)}", "F")

    # C: 金流科目（動態）
    for i, c in enumerate(categories, 1):
        lists.cell(row=i, column=3, value=c["name"])
    if categories:
        _dropdown(ws, f"_Lists!$C$1:$C${len(categories)}", "G")

    # D: 到期提醒天數
    for i in range(1, 8):
        lists.cell(row=i, column=4, value=i)
    _dropdown(ws, "_Lists!$D$1:$D$7", "H")

    # E: 是/否
    for i, v in enumerate(YN, 1):
        lists.cell(row=i, column=5, value=v)
    _dropdown(ws, "_Lists!$E$1:$E$2", "M")  # 出現叫貨系統
    _dropdown(ws, "_Lists!$E$1:$E$2", "N")  # 固定排程
    _dropdown(ws, "_Lists!$E$1:$E$2", "R")  # 國定假日休息

    return _to_bytes(wb)


# ─── 帳號範本 ─────────────────────────────────────────────────────────────────

ROLES = ["admin", "manager", "staff", "cashier"]

def build_accounts_template() -> bytes:
    """
    欄位：帳號*, 姓名*, 角色*, 密碼*, 零用金存取
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "帳號匯入"
    ws.freeze_panes = "A2"

    _header(ws, [
        ("帳號 *",    True,  16),
        ("姓名 *",    True,  14),
        ("角色 *",    True,  14),
        ("密碼 *",    True,  16),
        ("零用金存取", False, 14),
    ])
    _example(ws, ["john", "王大明", "staff", "pass123", "否"])

    lists = wb.create_sheet("_Lists")
    lists.sheet_state = "hidden"

    # A: 角色
    for i, r in enumerate(ROLES, 1):
        lists.cell(row=i, column=1, value=r)
    dv = DataValidation(
        type="list", formula1="_Lists!$A$1:$A$4",
        allowBlank=False, showErrorMessage=True,
        error="請填入：admin / manager / staff / cashier",
        errorTitle="無效角色"
    )
    ws.add_data_validation(dv)
    dv.add("C2:C2000")

    # B: 是/否
    for i, v in enumerate(YN, 1):
        lists.cell(row=i, column=2, value=v)
    _dropdown(ws, "_Lists!$B$1:$B$2", "E")

    return _to_bytes(wb)
